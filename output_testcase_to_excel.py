import textwrap
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def clean_text(text):
    return textwrap.dedent(text).strip()

test_info = {
    "test_name": "test_user_authentication",
    "description": """
    このテストは以下の機能を検証します：
    1. ユーザー認証の正常性
    2. データベース接続の確認
    3. エラーハンドリングの動作
    """,
    "sequence": """
    1. ユーザーIDとパスワードを入力
    2. 認証APIを呼び出し
    3. レスポンスを検証
    4. セッション情報を確認
    """,
    "expected": "認証成功とセッション作成"
}

# Excelワークブック作成
wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

# ヘッダー設定
headers = ["Test Name", "Description", "Sequence", "Expected"]
ws.append(headers)

# データ追加
ws.append([
    test_info["test_name"],
    clean_text(test_info["description"]),
    clean_text(test_info["sequence"]),
    test_info["expected"]
])

# セルの書式設定
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# 列幅の自動調整
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    
    # 最大80文字まで、最小15文字に制限
    adjusted_width = min(max(max_length, 15), 80)
    ws.column_dimensions[column_letter].width = adjusted_width

# 行の高さを自動調整
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    # 改行数に基づいて行の高さを調整
    max_lines = 1
    for cell in row:
        if cell.value:
            lines = str(cell.value).count('\n') + 1
            max_lines = max(max_lines, lines)
    
    ws.row_dimensions[row[0].row].height = max_lines * 15

# ファイル保存
wb.save('test_cases.xlsx')
print("Excelファイルを保存しました: test_cases.xlsx")
